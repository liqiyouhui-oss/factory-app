# -*- coding: utf-8 -*-
"""
================================================================
 工程管理アプリ Python+Excel版 バックエンドサーバー
================================================================
 
 役割: ブラウザから来たデータをExcelに書き、Excelから読んでブラウザに返す
 構成: Flask (Webサーバー) + openpyxl (Excel読み書き)
 起動: python app.py
 
 データファイル: factory_data.xlsx (このファイルと同じフォルダに自動生成)
 
 APIエンドポイント:
   GET  /                     → index.htmlを返す (画面表示)
   GET  /api/data/<key>       → 指定キーのデータをExcelから取得
   POST /api/data/<key>       → 指定キーのデータをExcelに保存
   GET  /api/data/_daily_list → 日報の全日付リストを取得
   
 【重要】 Excelファイルは排他ロック付きで安全に読み書きされる
 【重要】 全てのAPIでエラーハンドリング実装済み
================================================================
"""
 
import os
import json
import threading
import datetime
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory, abort
from openpyxl import Workbook, load_workbook
 
# ============================================================
# 基本設定
# ============================================================
 
# このファイルが置かれているフォルダのパス
BASE_DIR = Path(__file__).resolve().parent
# Excelデータファイルのパス
DATA_FILE = BASE_DIR / 'factory_data.xlsx'
# フロントエンドのHTMLが置かれているフォルダ
STATIC_DIR = BASE_DIR / 'static'
 
# Excelファイルの排他ロック（複数のリクエストが同時に書き込むと破損するため）
FILE_LOCK = threading.Lock()
 
# Flaskアプリ作成
app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path='')
 
 
# ============================================================
# シート定義
# ============================================================
# 各シート名とその列構成を定義
# 新しいシートを追加する場合はここに追加する
 
SHEETS = {
    # 受注一覧（顧客・製番・納期など）
    '受注一覧': [
        'id', 'seiban', 'client', 'product', 'productId', 'material', 'qty',
        'deadline', 'note',
        'heatTreat', 'heatTreatReturnDate',
        'outsource', 'outsourceReturnDate',
        'internalInspection', 'internalInspectionDate',
        'inspection', 'inspectionDate',
        'caution', 'cautionNote',
        'createdAt'
    ],
    # 工程明細（各受注の工程ステップ）
    # machineId: カンマ区切りで複数IDを保持可能（例: "1,2,3"）
    # hoursPerUnitUnit: 'h' または '日' （表示時の単位復元用）
    '工程明細': [
        'orderId', 'groupKey', 'stepIndex',
        'name', 'machineId', 'machineCount', 'hoursPerUnit', 'hoursPerUnitUnit',
        'done', 'actualHours',
        'scheduleOrder', 'scheduleStart', 'scheduleOperator'
    ],
    # 置き場在庫
    '置き場在庫': [
        'orderId', 'place', 'qty'
    ],
    # 製品マスタ
    '製品マスタ': [
        'id', 'name', 'category', 'note'
    ],
    # 製品×機械マスタ（どの製品をどの機械で加工できるか）
    '製品機械': [
        'productId', 'groupKey', 'machineId'
    ],
    # 機械マスタ
    '機械マスタ': [
        'id', 'name', 'type'
    ],
    # 作業者マスタ
    '作業者マスタ': [
        'id', 'name', 'team', 'role', 'mainMachineId', 'skills'
    ],
    # 勤務設定
    '勤務設定': [
        'key', 'value'
    ],
    # 日報メモ
    # 通常の連絡メモ: session, id(数値), text, checked を使用
    # 担当者別残業: session='_ot_', id=workerId(文字列), overtimeHours=残業時間
    '日報メモ': [
        'date', 'session', 'id', 'text', 'checked', 'overtimeHours'
    ],
    # カレンダー
    'カレンダー': [
        'id', 'date', 'title', 'type', 'color'
    ],
}
 
# フロントエンドのキーとExcelシートの対応表
# 既存フロントは「キー単位」でデータを取得するため、ここでシート⇔キーを変換
KEY_TO_SHEET = {
    'factory_py_v1': '受注一覧',          # 受注（工程明細と置き場在庫も一緒に返す）
    'factory_masters_v1': '機械マスタ',   # マスタ類（作業者・機械・勤務設定をまとめて返す）
    'factory_time_v1': None,              # 工数ヒント（メモリ保持。将来拡張用）
    'factory_cal_v1': 'カレンダー',       # カレンダーイベント
    'factory_products_v1': '製品マスタ',  # 製品マスタ
    'factory_product_machines_v1': '製品機械',  # 製品×機械
}
 
 
# ============================================================
# Excel初期化
# ============================================================
 
def init_excel():
    """
    初回起動時にExcelファイルを作成する。
    既存ファイルがある場合は不足シートのみ追加する。
    """
    # ファイルが無い場合は新規作成
    if not DATA_FILE.exists():
        wb = Workbook()
        # デフォルトシートを削除
        default_sheet = wb.active
        wb.remove(default_sheet)
        # 全シートを作成
        for sheet_name, columns in SHEETS.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(columns)  # ヘッダー行
        # 勤務設定のデフォルト値を投入
        ws_work = wb['勤務設定']
        ws_work.append(['startTime', '08:00'])
        ws_work.append(['endTime', '17:00'])
        ws_work.append(['breaks', json.dumps([
            {'id': 1, 'start': '10:00', 'end': '10:10'},
            {'id': 2, 'start': '15:00', 'end': '15:10'}
        ], ensure_ascii=False)])
        wb.save(str(DATA_FILE))
        wb.close()
        print(f'[init] Excelファイルを新規作成しました: {DATA_FILE}')
        return
 
    # 既存ファイルがある場合は不足シートを追加
    try:
        wb = load_workbook(str(DATA_FILE))
        modified = False
        for sheet_name, columns in SHEETS.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(columns)
                modified = True
                print(f'[init] 不足シートを追加: {sheet_name}')
        if modified:
            wb.save(str(DATA_FILE))
        wb.close()
    except Exception as e:
        print(f'[init] エラー: {e}')
        raise
 
 
# ============================================================
# Excel読み書きヘルパー
# ============================================================
 
def read_sheet(sheet_name):
    """
    指定シートを全行読み込んで辞書のリストで返す。
    例: [{'id':1, 'name':'山田'}, {'id':2, 'name':'佐藤'}]
    """
    if not DATA_FILE.exists():
        init_excel()
 
    with FILE_LOCK:
        try:
            wb = load_workbook(str(DATA_FILE), read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return []
            ws = wb[sheet_name]
            # ヘッダー行を取得
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return []
            headers = list(rows[0])
            # データ行を辞書化
            result = []
            for row in rows[1:]:
                # 全てNoneの空行はスキップ
                if all(v is None for v in row):
                    continue
                record = {}
                for i, h in enumerate(headers):
                    if h is None:
                        continue
                    v = row[i] if i < len(row) else None
                    # Excelのブール値や日付を適切に変換
                    if isinstance(v, datetime.datetime):
                        v = v.strftime('%Y-%m-%d')
                    elif isinstance(v, datetime.date):
                        v = v.strftime('%Y-%m-%d')
                    record[h] = v
                result.append(record)
            return result
        except Exception as e:
            print(f'[read_sheet] エラー ({sheet_name}): {e}')
            return []
 
 
def write_sheet(sheet_name, records):
    """
    指定シートを丸ごと書き換える。
    recordsは辞書のリスト。
    """
    if not DATA_FILE.exists():
        init_excel()
 
    with FILE_LOCK:
        try:
            wb = load_workbook(str(DATA_FILE))
            # シートが無ければ作る
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                columns = SHEETS.get(sheet_name, [])
                if columns:
                    ws.append(columns)
            else:
                ws = wb[sheet_name]
            # 既存データをクリア（ヘッダーは残す）
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            # ヘッダーを取得
            headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else SHEETS.get(sheet_name, [])
            if not headers:
                headers = SHEETS.get(sheet_name, [])
                if headers:
                    ws.append(headers)
            # データを追加
            for rec in records:
                if not isinstance(rec, dict):
                    continue
                row = []
                for h in headers:
                    v = rec.get(h)
                    # リストや辞書はJSON文字列化して保存
                    if isinstance(v, (list, dict)):
                        v = json.dumps(v, ensure_ascii=False)
                    row.append(v)
                ws.append(row)
            wb.save(str(DATA_FILE))
            wb.close()
            return True
        except Exception as e:
            print(f'[write_sheet] エラー ({sheet_name}): {e}')
            return False
 
 
# ============================================================
# メモリキャッシュ（Excelを使わない一時データ）
# ============================================================
 
# 日報メモは日付ごとにシート分けせず、全日付を1シートに保存
# メモリキャッシュで高速化
_memory_cache = {}
 
 
# ============================================================
# 統合データ取得（既存フロント互換）
# ============================================================
 
def get_orders():
    """
    受注データを取得する。
    工程明細と置き場在庫を結合して、既存フロントが期待する形にする。
    """
    orders = read_sheet('受注一覧')
    steps_all = read_sheet('工程明細')
    storages_all = read_sheet('置き場在庫')
 
    result = []
    for o in orders:
        order_id = o.get('id')
        # ID型をint化して比較（Excelから読み込むと文字列になることがある）
        try:
            oid_int = int(order_id) if order_id is not None else None
        except (ValueError, TypeError):
            oid_int = None
 
        # 該当の工程明細を取得
        group1 = []
        group2 = []
        for s in steps_all:
            try:
                sid_int = int(s.get('orderId')) if s.get('orderId') is not None else None
            except (ValueError, TypeError):
                sid_int = None
            if sid_int != oid_int:
                continue
            step_data = {
                'id': int(s.get('stepIndex') or 0),
                'name': s.get('name') or '',
                'machineId': s.get('machineId') or '',
                'operatorId': '',  # 旧互換用
                'machineCount': int(s.get('machineCount') or 1),
                'hoursPerUnit': s.get('hoursPerUnit') or '',
                'hoursPerUnitUnit': s.get('hoursPerUnitUnit') or 'h',
                'done': int(s.get('done') or 0),
                'actualHours': s.get('actualHours') or '',
                'scheduleOrder': s.get('scheduleOrder'),
                'scheduleStart': s.get('scheduleStart') or '',
                'scheduleOperator': s.get('scheduleOperator') or '',
            }
            if s.get('groupKey') == 'group1':
                group1.append(step_data)
            elif s.get('groupKey') == 'group2':
                group2.append(step_data)
        # 順序でソート
        group1.sort(key=lambda x: x['id'])
        group2.sort(key=lambda x: x['id'])
 
        # 置き場在庫
        storage_qty = {}
        for st in storages_all:
            try:
                sid_int = int(st.get('orderId')) if st.get('orderId') is not None else None
            except (ValueError, TypeError):
                sid_int = None
            if sid_int == oid_int:
                place = st.get('place') or ''
                qty = int(st.get('qty') or 0)
                if place:
                    storage_qty[place] = qty
 
        # 元の受注データをコピー
        order_data = dict(o)
        order_data['group1'] = group1 if group1 else []
        order_data['group2'] = group2 if group2 else []
        order_data['storageQty'] = storage_qty
        # ブール値の正規化
        for bk in ['heatTreat', 'outsource', 'inspection', 'internalInspection', 'caution']:
            v = order_data.get(bk)
            order_data[bk] = v in (True, 'TRUE', 'true', 'True', 1, '1')
        result.append(order_data)
 
    return result
 
 
def save_orders(orders):
    """
    受注データを保存する。
    受注一覧・工程明細・置き場在庫の3シートを更新する。
    """
    order_records = []
    step_records = []
    storage_records = []
 
    for o in orders:
        if not isinstance(o, dict):
            continue
        # 受注本体
        order_rec = {
            'id': o.get('id'),
            'seiban': o.get('seiban') or '',
            'client': o.get('client') or '',
            'product': o.get('product') or '',
            'productId': o.get('productId') or '',
            'material': o.get('material') or '',
            'qty': o.get('qty') or 0,
            'deadline': o.get('deadline') or '',
            'note': o.get('note') or '',
            'heatTreat': bool(o.get('heatTreat')),
            'heatTreatReturnDate': o.get('heatTreatReturnDate') or '',
            'outsource': bool(o.get('outsource')),
            'outsourceReturnDate': o.get('outsourceReturnDate') or '',
            'internalInspection': bool(o.get('internalInspection')),
            'internalInspectionDate': o.get('internalInspectionDate') or '',
            'inspection': bool(o.get('inspection')),
            'inspectionDate': o.get('inspectionDate') or '',
            'caution': bool(o.get('caution')),
            'cautionNote': o.get('cautionNote') or '',
            'createdAt': o.get('at') or o.get('createdAt') or '',
        }
        order_records.append(order_rec)
 
        # 工程明細
        for gk in ['group1', 'group2']:
            steps = o.get(gk) or []
            for idx, s in enumerate(steps):
                if not isinstance(s, dict):
                    continue
                step_rec = {
                    'orderId': o.get('id'),
                    'groupKey': gk,
                    'stepIndex': idx,
                    'name': s.get('name') or '',
                    'machineId': s.get('machineId') or '',
                    'machineCount': s.get('machineCount') or 1,
                    'hoursPerUnit': s.get('hoursPerUnit') or '',
                    'hoursPerUnitUnit': s.get('hoursPerUnitUnit') or 'h',
                    'done': s.get('done') or 0,
                    'actualHours': s.get('actualHours') or '',
                    'scheduleOrder': s.get('scheduleOrder'),
                    'scheduleStart': s.get('scheduleStart') or '',
                    'scheduleOperator': s.get('scheduleOperator') or '',
                }
                step_records.append(step_rec)
 
        # 置き場在庫
        storage_qty = o.get('storageQty') or {}
        if isinstance(storage_qty, dict):
            for place, qty in storage_qty.items():
                if qty and int(qty) > 0:
                    storage_records.append({
                        'orderId': o.get('id'),
                        'place': place,
                        'qty': int(qty)
                    })
 
    # 3シート全更新
    ok1 = write_sheet('受注一覧', order_records)
    ok2 = write_sheet('工程明細', step_records)
    ok3 = write_sheet('置き場在庫', storage_records)
    return ok1 and ok2 and ok3
 
 
def get_masters():
    """
    マスタ類をまとめて取得する（既存フロント互換）
    """
    workers = read_sheet('作業者マスタ')
    machines = read_sheet('機械マスタ')
    work_settings_rows = read_sheet('勤務設定')
 
    # 勤務設定をdict化
    ws_dict = {
        'startTime': '08:00',
        'endTime': '17:00',
        'breaks': [
            {'id': 1, 'start': '10:00', 'end': '10:10'},
            {'id': 2, 'start': '15:00', 'end': '15:10'}
        ],
    }
    for row in work_settings_rows:
        k = row.get('key')
        v = row.get('value')
        if k == 'breaks' and v:
            try:
                ws_dict['breaks'] = json.loads(v) if isinstance(v, str) else v
            except (json.JSONDecodeError, TypeError):
                pass
        elif k in ('startTime', 'endTime'):
            ws_dict[k] = v or ws_dict[k]
 
    # workersのskillsをJSON→listに変換
    normalized_workers = []
    for w in workers:
        wd = dict(w)
        skills = wd.get('skills')
        if isinstance(skills, str) and skills:
            try:
                wd['skills'] = json.loads(skills)
            except (json.JSONDecodeError, TypeError):
                wd['skills'] = []
        elif not skills:
            wd['skills'] = []
        normalized_workers.append(wd)
 
    return {
        'workers': normalized_workers,
        'machines': machines,
        'workSettings': ws_dict,
    }
 
 
def save_masters(data):
    """
    マスタ類をまとめて保存する
    """
    if not isinstance(data, dict):
        return False
 
    workers = data.get('workers') or []
    machines = data.get('machines') or []
    ws = data.get('workSettings') or {}
 
    # 勤務設定をkey-value形式に変換
    ws_rows = [
        {'key': 'startTime', 'value': ws.get('startTime', '08:00')},
        {'key': 'endTime', 'value': ws.get('endTime', '17:00')},
        {'key': 'breaks', 'value': json.dumps(ws.get('breaks', []), ensure_ascii=False)},
    ]
 
    ok1 = write_sheet('作業者マスタ', workers)
    ok2 = write_sheet('機械マスタ', machines)
    ok3 = write_sheet('勤務設定', ws_rows)
    return ok1 and ok2 and ok3
 
 
# ============================================================
# APIエンドポイント
# ============================================================
 
@app.route('/')
def index():
    """トップページ（index.htmlを返す）"""
    index_path = STATIC_DIR / 'index.html'
    if not index_path.exists():
        return '<h1>index.htmlが見つかりません</h1><p>staticフォルダにindex.htmlを配置してください。</p>', 404
    # キャッシュ無効化（アプリ更新時に古いバージョンが表示されないように）
    response = send_from_directory(str(STATIC_DIR), 'index.html')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response
 
 
@app.route('/api/data/<path:key>', methods=['GET'])
def api_get(key):
    """データ取得API"""
    try:
        # 日報メモ（キーに日付が含まれる）
        if key.startswith('factory_daily_v1_'):
            date = key.replace('factory_daily_v1_', '')
            all_daily = read_sheet('日報メモ')
            meetings_rows = [m for m in all_daily if m.get('date') == date]
            # フロントが期待する形式に整形
            # 通常のmeeting行と、担当者別残業行(session='_ot_')を分離
            formatted = []
            overtime_by_worker = {}
            for m in meetings_rows:
                session = m.get('session') or ''
                if session == '_ot_':
                    # 担当者別残業レコード: id=workerId, overtimeHours=時間
                    try:
                        wid = str(m.get('id') or '')
                        hours = float(m.get('overtimeHours') or 0)
                        if wid and hours > 0:
                            overtime_by_worker[wid] = hours
                    except (ValueError, TypeError):
                        pass
                    continue
                # 通常のミーティングメモ
                try:
                    mid = int(m.get('id') or 0)
                except (ValueError, TypeError):
                    mid = 0
                formatted.append({
                    'id': mid,
                    'session': session,
                    'text': m.get('text') or '',
                    'checked': m.get('checked') in (True, 'TRUE', 'true', 'True', 1, '1'),
                })
            return jsonify({'ok': True, 'value': {
                'meetings': formatted,
                'overtimeByWorker': overtime_by_worker
            }})
 
        # 受注
        if key == 'factory_py_v1':
            orders = get_orders()
            return jsonify({'ok': True, 'value': orders})
 
        # マスタ
        if key == 'factory_masters_v1':
            masters = get_masters()
            return jsonify({'ok': True, 'value': masters})
 
        # カレンダー
        if key == 'factory_cal_v1':
            events = read_sheet('カレンダー')
            return jsonify({'ok': True, 'value': events})
 
        # 製品マスタ
        if key == 'factory_products_v1':
            products = read_sheet('製品マスタ')
            return jsonify({'ok': True, 'value': products})
 
        # 製品×機械マスタ
        if key == 'factory_product_machines_v1':
            pm = read_sheet('製品機械')
            return jsonify({'ok': True, 'value': pm})
 
        # 工数ヒント（メモリ保持）
        if key == 'factory_time_v1':
            return jsonify({'ok': True, 'value': _memory_cache.get(key, {})})
 
        # 未知のキー
        return jsonify({'ok': True, 'value': None})
 
    except Exception as e:
        print(f'[api_get] エラー ({key}): {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500
 
 
@app.route('/api/data/<path:key>', methods=['POST'])
def api_post(key):
    """データ保存API"""
    try:
        body = request.get_json(silent=True)
        if body is None:
            return jsonify({'ok': False, 'error': 'JSONボディが必要'}), 400
        value = body.get('value')
 
        # 日報メモ
        if key.startswith('factory_daily_v1_'):
            date = key.replace('factory_daily_v1_', '')
            # 既存の全日付データを取得
            all_daily = read_sheet('日報メモ')
            # 該当日付を削除
            all_daily = [m for m in all_daily if m.get('date') != date]
            # 新しいデータを追加
            if isinstance(value, dict):
                meetings = value.get('meetings') or []
                # 通常のミーティングメモを追加
                for m in meetings:
                    if not isinstance(m, dict):
                        continue
                    all_daily.append({
                        'date': date,
                        'session': m.get('session') or '',
                        'id': m.get('id'),
                        'text': m.get('text') or '',
                        'checked': bool(m.get('checked')),
                        'overtimeHours': '',
                    })
                # 担当者別残業を保存 (session='_ot_' で区別)
                ot_map = value.get('overtimeByWorker') or {}
                if isinstance(ot_map, dict):
                    for worker_id, hours in ot_map.items():
                        try:
                            h = float(hours)
                            if h <= 0:
                                continue
                        except (ValueError, TypeError):
                            continue
                        all_daily.append({
                            'date': date,
                            'session': '_ot_',
                            'id': str(worker_id),
                            'text': '',
                            'checked': False,
                            'overtimeHours': h,
                        })
            ok = write_sheet('日報メモ', all_daily)
            return jsonify({'ok': ok})
 
        # 受注
        if key == 'factory_py_v1':
            if not isinstance(value, list):
                return jsonify({'ok': False, 'error': '配列が必要'}), 400
            ok = save_orders(value)
            return jsonify({'ok': ok})
 
        # マスタ
        if key == 'factory_masters_v1':
            ok = save_masters(value)
            return jsonify({'ok': ok})
 
        # カレンダー
        if key == 'factory_cal_v1':
            if not isinstance(value, list):
                return jsonify({'ok': False, 'error': '配列が必要'}), 400
            ok = write_sheet('カレンダー', value)
            return jsonify({'ok': ok})
 
        # 製品マスタ
        if key == 'factory_products_v1':
            if not isinstance(value, list):
                return jsonify({'ok': False, 'error': '配列が必要'}), 400
            ok = write_sheet('製品マスタ', value)
            return jsonify({'ok': ok})
 
        # 製品×機械マスタ
        if key == 'factory_product_machines_v1':
            if not isinstance(value, list):
                return jsonify({'ok': False, 'error': '配列が必要'}), 400
            ok = write_sheet('製品機械', value)
            return jsonify({'ok': ok})
 
        # 工数ヒント（メモリ）
        if key == 'factory_time_v1':
            _memory_cache[key] = value or {}
            return jsonify({'ok': True})
 
        # 未知のキー
        return jsonify({'ok': False, 'error': f'未知のキー: {key}'}), 400
 
    except Exception as e:
        print(f'[api_post] エラー ({key}): {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500
 
 
@app.errorhandler(404)
def not_found(e):
    return jsonify({'ok': False, 'error': 'Not Found'}), 404
 
 
@app.errorhandler(500)
def server_error(e):
    return jsonify({'ok': False, 'error': 'Internal Server Error'}), 500
 
 
# ============================================================
# サーバー起動
# ============================================================
 
if __name__ == '__main__':
    print('=' * 60)
    print(' 工程管理アプリ Python+Excel版 サーバー起動')
    print('=' * 60)
    print(f' データファイル: {DATA_FILE}')
    print(f' 静的ファイル: {STATIC_DIR}')
 
    # Excelファイルを初期化
    try:
        init_excel()
        print(' Excelファイル初期化: OK')
    except Exception as e:
        print(f' Excelファイル初期化エラー: {e}')
        raise SystemExit(1)
 
    print('=' * 60)
    print(' ブラウザで http://localhost:5000 にアクセス')
    print(' 他のPC/スマホからは http://<このPCのIP>:5000')
    print(' 停止するには Ctrl+C を押してください')
    print('=' * 60)
 
    # 0.0.0.0 = LAN内の他の機器からもアクセス可能にする
    # debug=False = 本番運用モード（エラー詳細を外部に見せない）
    app.run(host='0.0.0.0', port=5000, debug=False)
